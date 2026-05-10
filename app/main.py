from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List
from datetime import datetime
from urllib.parse import quote  # NEW: for URL-safe filenames

import os

from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles

from app.services.pdf_extractor import extract_fields_from_pdf
from app.services.batch_processor import process_pdf_batch
from app.services.employee_db import EmployeeDB
from app.services.id_suggest import suggest_ids

import openpyxl


app = FastAPI(title="PDF Report Analyzer")

# --------------------------------------------------
# Paths
# --------------------------------------------------
_DEFAULT_SHARED_ROOT = Path(r"I:\60 - Services\30 - BI\010 - Shared\HP_app")
SHARED_ROOT = Path(os.getenv("HP_APP_ROOT", str(_DEFAULT_SHARED_ROOT)))

INPUT_DIR = Path(os.getenv("HP_INPUT_DIR", str(SHARED_ROOT / "input_pdfs")))
EXPORTS_ROOT = Path(os.getenv("HP_EXPORTS_ROOT", str(SHARED_ROOT / "exports")))

UPLOAD_DIR = Path("uploaded_pdfs")  # local uploads (optional)
UPLOAD_DIR.mkdir(exist_ok=True)

# Where /review looks for batches (subfolders named export_*)
print(
    f"[PATHS] EXPORTS_ROOT={EXPORTS_ROOT} "
    f"(exists={EXPORTS_ROOT.exists()}) — batches = child dirs named export_*"
)
print(f"[PATHS] INPUT_DIR={INPUT_DIR} (exists={INPUT_DIR.exists()})")

# --------------------------------------------------
# Templates (UI)
# --------------------------------------------------
templates = Jinja2Templates(directory="app/templates")

# --------------------------------------------------
# Database (for suggestions)
# --------------------------------------------------
DB = EmployeeDB(str(EXPORTS_ROOT / "employees.xlsx"))  # keep in exports as you do


# ==================================================
# Helpers
# ==================================================
def _load_manifest(export_dir: Path) -> Dict[str, Any]:
    manifest_path = export_dir / "manifest.json"
    if not manifest_path.exists():
        raise HTTPException(status_code=404, detail=f"manifest.json not found: {manifest_path}")

    try:
        return json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to read manifest.json: {e}")


def _docs_root(export_dir: Path) -> Path:
    p = export_dir / "docs"
    if not p.exists():
        raise HTTPException(status_code=404, detail=f"docs folder not found: {p}")
    return p


def _safe_export_folder(export_name: str) -> Path:
    if any(x in export_name for x in ("/", "\\", "..")):
        raise HTTPException(status_code=400, detail="Invalid export_name")

    export_dir = EXPORTS_ROOT / export_name
    if not export_dir.exists() or not export_dir.is_dir():
        raise HTTPException(status_code=404, detail=f"Export folder not found: {export_dir}")
    return export_dir


def _safe_doc_folder(docs_root: Path, doc_name: str) -> Path:
    if any(x in doc_name for x in ("/", "\\", "..")):
        raise HTTPException(status_code=400, detail="Invalid doc_name")

    doc_dir = docs_root / doc_name
    if not doc_dir.exists() or not doc_dir.is_dir():
        raise HTTPException(status_code=404, detail=f"Document folder not found: {doc_dir}")
    return doc_dir


def _coerce_doc_to_name(d: Any) -> str:
    if isinstance(d, str):
        return d
    if isinstance(d, dict):
        for key in ("doc_name", "name", "folder_name", "id", "filename"):
            val = d.get(key)
            if isinstance(val, str) and val.strip():
                return val.strip()
    return str(d)


def _list_docs_from_manifest(manifest: Dict[str, Any], docs_root: Path) -> List[str]:
    docs: List[str] = []

    if isinstance(manifest.get("documents"), list):
        for item in manifest["documents"]:
            name = _coerce_doc_to_name(item)
            if name and name != "[object Object]":
                docs.append(name)

    elif isinstance(manifest.get("docs"), list):
        for item in manifest["docs"]:
            name = _coerce_doc_to_name(item)
            if name and name != "[object Object]":
                docs.append(name)

    elif isinstance(manifest.get("items"), list):
        for item in manifest["items"]:
            name = _coerce_doc_to_name(item)
            if name and name != "[object Object]":
                docs.append(name)

    if not docs:
        docs = sorted([p.name for p in docs_root.iterdir() if p.is_dir()])

    # de-dup preserve order
    seen = set()
    out: List[str] = []
    for x in docs:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


# ---------------------------
# per-doc review storage
# ---------------------------
def _review_path(export_dir: Path, doc_name: str) -> Path:
    return export_dir / "docs" / doc_name / "review.json"


def _load_review(export_dir: Path, doc_name: str) -> Dict[str, Any]:
    p = _review_path(export_dir, doc_name)
    if not p.exists():
        return {}
    try:
        return json.loads(p.read_text(encoding="utf-8"))
    except Exception:
        return {}


def _merge_review_into_results(results: Dict[str, Any], review: Dict[str, Any]) -> Dict[str, Any]:
    if not review or "rows" not in review:
        return results

    review_map: Dict[int, Dict[str, Any]] = {}
    for rr in review.get("rows", []):
        try:
            review_map[int(rr.get("row"))] = rr
        except Exception:
            continue

    rows = results.get("rows", [])
    if isinstance(rows, list):
        for r in rows:
            try:
                row_idx = int(r.get("row"))
            except Exception:
                continue
            if row_idx in review_map:
                saved = review_map[row_idx]
                if "resolved_id" in saved:
                    r["resolved_id"] = saved.get("resolved_id")
                if "resolved_name" in saved:
                    r["resolved_name"] = saved.get("resolved_name")

    results["_review_saved_at"] = review.get("saved_at")
    return results


# ---------------------------
# submit-once log
# ---------------------------
SUBMIT_LOG_PATH = EXPORTS_ROOT / "submitted_docs.xlsx"


def _ensure_submit_log():
    if not SUBMIT_LOG_PATH.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "log"
        ws.append(["Export", "Document", "Submitted_At"])
        wb.save(SUBMIT_LOG_PATH)


def _is_doc_already_submitted(export_name: str, doc_name: str) -> bool:
    _ensure_submit_log()
    wb = openpyxl.load_workbook(SUBMIT_LOG_PATH)
    ws = wb["log"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        e = str(r[0] or "").strip()
        d = str(r[1] or "").strip()
        if e == export_name and d == doc_name:
            return True
    return False


def _mark_doc_submitted(export_name: str, doc_name: str):
    _ensure_submit_log()
    wb = openpyxl.load_workbook(SUBMIT_LOG_PATH)
    ws = wb["log"]
    ws.append([export_name, doc_name, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    wb.save(SUBMIT_LOG_PATH)


# ---------------------------
# PDF link helpers
# ---------------------------
def _get_manifest_source_pdf(export_dir: Path, doc_name: str) -> str | None:
    """
    Find the original PDF filename for a given doc folder using manifest.json.
    Looks for keys: source_pdf / pdf / source.
    """
    manifest_path = export_dir / "manifest.json"
    if not manifest_path.exists():
        return None

    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except Exception:
        return None

    docs: Any = None
    if isinstance(manifest.get("documents"), list):
        docs = manifest["documents"]
    elif isinstance(manifest.get("docs"), list):
        docs = manifest["docs"]
    elif isinstance(manifest.get("items"), list):
        docs = manifest["items"]

    if not isinstance(docs, list):
        return None

    for item in docs:
        if isinstance(item, dict):
            dn = (item.get("doc_name") or item.get("name") or item.get("folder_name") or "").strip()
            sp = (item.get("source_pdf") or item.get("pdf") or item.get("source") or "").strip()
            if dn == doc_name and sp:
                return sp

    return None


def _extract_sa_ids(s: str) -> List[str]:
    # matches SA12345 and SA012345 etc
    return sorted(set(re.findall(r"\bSA\d+\b", s, flags=re.IGNORECASE)))


def _score_pdf_match(doc_name: str, pdf_filename: str) -> int:
    """
    Score based on how many SA IDs are common.
    If no SA IDs exist, fall back to token overlap.
    """
    doc_ids = set(_extract_sa_ids(doc_name))
    pdf_ids = set(_extract_sa_ids(pdf_filename))

    if doc_ids:
        return len(doc_ids.intersection(pdf_ids))

    # fallback: token overlap (very rough)
    def norm_tokens(x: str) -> set[str]:
        x = x.lower()
        x = re.sub(r"[^a-z0-9]+", " ", x)
        toks = {t for t in x.split() if len(t) >= 3}
        return toks

    return len(norm_tokens(doc_name).intersection(norm_tokens(pdf_filename)))


def _find_pdf_for_doc(export_dir: Path, doc_name: str) -> str | None:
    """
    Return a PDF filename that exists in INPUT_DIR, for this doc.
    Priority:
      1) manifest source_pdf
      2) exact match (doc_name.pdf)
      3) best match by SA IDs overlap (most reliable)
      4) contains match (last resort)
    """
    if not INPUT_DIR.exists():
        return None

    # 1) Manifest
    sp = _get_manifest_source_pdf(export_dir, doc_name)
    if sp and (INPUT_DIR / sp).exists():
        return sp

    # 2) Exact doc_name + extension
    cand = list(INPUT_DIR.glob(f"{doc_name}.pdf")) + list(INPUT_DIR.glob(f"{doc_name}.PDF"))
    if cand:
        return cand[0].name

    # 3) Best SA-ID overlap score
    pdfs = list(INPUT_DIR.glob("*.pdf")) + list(INPUT_DIR.glob("*.PDF"))
    if pdfs:
        scored = [(p.name, _score_pdf_match(doc_name, p.name)) for p in pdfs]
        scored.sort(key=lambda x: x[1], reverse=True)
        best_name, best_score = scored[0]
        # require at least 1 ID overlap if doc has SA IDs
        if _extract_sa_ids(doc_name):
            if best_score >= 1:
                return best_name
        else:
            # if no SA ids in doc_name, accept weak score threshold
            if best_score >= 2:
                return best_name

    # 4) contains match
    lowered = doc_name.lower()
    for p in pdfs:
        if lowered in p.name.lower():
            return p.name

    return None


# ==================================================
# Static mounts
# ==================================================
if EXPORTS_ROOT.exists():
    app.mount("/static", StaticFiles(directory=str(EXPORTS_ROOT)), name="static")
    print(f"[STATIC] Mounted /static -> {EXPORTS_ROOT}")
else:
    print(f"[WARN] EXPORTS_ROOT does not exist yet: {EXPORTS_ROOT}")

if INPUT_DIR.exists():
    app.mount("/pdfs", StaticFiles(directory=str(INPUT_DIR)), name="pdfs")
    print(f"[PDFS] Mounted /pdfs -> {INPUT_DIR}")
else:
    print(f"[WARN] INPUT_DIR does not exist yet: {INPUT_DIR}")


# ==================================================
# Health
# ==================================================
@app.get("/health")
def health_check():
    return {"status": "ok"}


# ==================================================
# Review UI
# ==================================================
@app.get("/review", response_class=HTMLResponse)
async def review_page(request: Request):
    return templates.TemplateResponse("review.html", {"request": request})


# ==================================================
# API: list batches
# ==================================================
def _is_complete_export_batch(p: Path) -> bool:
    """Ignore partial/crashed folders (e.g. OCR failed mid-run)."""
    return p.is_dir() and (p / "manifest.json").is_file() and (p / "docs").is_dir()


@app.get("/api/batches")
def api_batches():
    if not EXPORTS_ROOT.exists():
        return JSONResponse({"batches": []})

    batches = sorted(
        [p for p in EXPORTS_ROOT.iterdir() if _is_complete_export_batch(p) and p.name.startswith("export_")],
        key=lambda p: p.name,
        reverse=True,
    )

    return JSONResponse({"batches": [{"id": b.name, "title": b.name} for b in batches]})


# ==================================================
# API: list docs in batch
# ==================================================
@app.get("/api/batches/{export_name}/docs")
def api_batch_docs(export_name: str):
    export_dir = _safe_export_folder(export_name)
    manifest = _load_manifest(export_dir)
    docs_root = _docs_root(export_dir)

    docs = _list_docs_from_manifest(manifest, docs_root)
    return JSONResponse({"export": export_dir.name, "documents": docs})


# ==================================================
# API: read one doc results.json (MERGED with review.json)
# ==================================================
@app.get("/api/batches/{export_name}/docs/{doc_name}")
def api_batch_doc_detail(export_name: str, doc_name: str):
    export_dir = _safe_export_folder(export_name)
    docs_root = _docs_root(export_dir)
    doc_dir = _safe_doc_folder(docs_root, doc_name)

    results_path = doc_dir / "results.json"
    if not results_path.exists():
        raise HTTPException(status_code=404, detail=f"results.json not found: {results_path}")

    try:
        data = json.loads(results_path.read_text(encoding="utf-8"))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed reading results.json: {e}")

    data["_export"] = export_dir.name
    data["_doc_name"] = doc_name
    data["_static_base"] = f"/static/{export_dir.name}/docs/{doc_name}"

    # PDF link (robust)
    pdf_name = _find_pdf_for_doc(export_dir, doc_name)
    if pdf_name:
        data["_pdf_available"] = True
        # IMPORTANT: URL-safe encoding for spaces/() etc
        data["_pdf_url"] = f"/pdfs/{quote(pdf_name)}"
        data["_pdf_name"] = pdf_name
    else:
        data["_pdf_available"] = False
        data["_pdf_url"] = None
        data["_pdf_name"] = None

    review = _load_review(export_dir, doc_name)
    data = _merge_review_into_results(data, review)

    return JSONResponse(data)


# ==================================================
# API: explicit pdf-link endpoint (optional)
# ==================================================
@app.get("/api/batches/{export_name}/docs/{doc_name}/pdf-link")
def api_doc_pdf_link(export_name: str, doc_name: str):
    export_dir = _safe_export_folder(export_name)
    pdf_name = _find_pdf_for_doc(export_dir, doc_name)
    if not pdf_name:
        return {"available": False, "pdf_url": None, "pdf_name": None}
    return {"available": True, "pdf_url": f"/pdfs/{quote(pdf_name)}", "pdf_name": pdf_name}


# ==================================================
# Save per-doc review (writes review.json)
# ==================================================
@app.post("/api/batches/{export_name}/docs/{doc_name}/save-review")
async def save_review_for_doc(export_name: str, doc_name: str, payload: Dict[str, Any]):
    export_dir = _safe_export_folder(export_name)
    docs_root = _docs_root(export_dir)
    _ = _safe_doc_folder(docs_root, doc_name)

    rows = payload.get("rows", [])
    if not isinstance(rows, list):
        raise HTTPException(status_code=400, detail="rows must be a list")

    cleaned: List[Dict[str, Any]] = []
    for r in rows:
        try:
            row_idx = int(r.get("row"))
        except Exception:
            continue
        cleaned.append({
            "row": row_idx,
            "resolved_id": str(r.get("resolved_id") or "").strip(),
            "resolved_name": str(r.get("resolved_name") or "").strip(),
        })

    review_doc = {
        "export": export_name,
        "doc": doc_name,
        "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "rows": cleaned
    }

    p = _review_path(export_dir, doc_name)
    p.write_text(json.dumps(review_doc, ensure_ascii=False, indent=2), encoding="utf-8")

    return {"status": "ok", "review_file": str(p)}


# ==================================================
# Suggest IDs (returns id + name + score)
# ==================================================
@app.get("/suggest-ids")
async def suggest_ids_endpoint(q: str, department: str = "CASTHOUSE", limit: int = 10):
    dept_records = DB.get_records_for_department(department)

    valid_ids: List[str] = []
    id_to_name: Dict[str, str] = {}

    for rec in dept_records:
        if isinstance(rec, tuple):
            emp_id = rec[0]
            emp_name = rec[1] if len(rec) > 1 else ""
        else:
            emp_id = rec.get("emp_id")
            emp_name = rec.get("name") or rec.get("emp_name") or rec.get("full_name") or ""

        if emp_id:
            sid = str(emp_id)
            valid_ids.append(sid)
            if emp_name:
                id_to_name[sid] = str(emp_name)

    suggestions = suggest_ids(q, valid_ids, limit=limit)

    return JSONResponse({
        "suggestions": [
            {"id": sid, "name": id_to_name.get(sid, ""), "score": score}
            for sid, score in suggestions
        ]
    })


# ==================================================
# Submit summary -> review_summary.xlsx (submit once per doc)
# ==================================================
@app.post("/api/submit-summary")
async def submit_summary(payload: Dict[str, Any]):
    export_name = payload.get("export")
    doc_name = payload.get("doc")
    rows = payload.get("rows", [])

    if not export_name or not doc_name or not isinstance(rows, list):
        raise HTTPException(status_code=400, detail="Invalid payload")

    if _is_doc_already_submitted(export_name, doc_name):
        return {
            "status": "ok",
            "already_submitted": True,
            "document": doc_name,
            "export": export_name
        }

    out_path = EXPORTS_ROOT / "review_summary.xlsx"

    if not out_path.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "summary"
        ws.append(["ID", "Name", "Count", "Documents"])
        wb.save(out_path)

    wb = openpyxl.load_workbook(out_path)
    ws = wb["summary"]

    existing: Dict[str, Dict[str, Any]] = {}
    for r in ws.iter_rows(min_row=2, values_only=False):
        emp_id_cell, name_cell, count_cell, docs_cell = r[0], r[1], r[2], r[3]
        emp_id = str(emp_id_cell.value).strip() if emp_id_cell.value is not None else ""
        if not emp_id:
            continue

        name = str(name_cell.value).strip() if name_cell.value else ""
        count = int(count_cell.value) if count_cell.value not in (None, "") else 0
        docs_raw = str(docs_cell.value).strip() if docs_cell.value else ""
        docs_set = set([d.strip() for d in docs_raw.split(",") if d.strip()])

        existing[emp_id] = {
            "row": emp_id_cell.row,
            "name": name,
            "count": count,
            "docs": docs_set,
        }

    doc_counts: Dict[str, Dict[str, Any]] = {}
    for rr in rows:
        emp_id = str(rr.get("resolved_id") or "").strip()
        emp_name = str(rr.get("resolved_name") or "").strip()
        if not emp_id or not emp_name:
            continue

        if emp_id not in doc_counts:
            doc_counts[emp_id] = {"name": emp_name, "count": 0}
        doc_counts[emp_id]["count"] += 1

        if len(emp_name) > len(doc_counts[emp_id]["name"]):
            doc_counts[emp_id]["name"] = emp_name

    for emp_id, info in doc_counts.items():
        add_count = int(info["count"])
        new_name = str(info["name"])

        if emp_id in existing:
            row_idx = existing[emp_id]["row"]
            old_count = existing[emp_id]["count"]
            ws.cell(row=row_idx, column=3).value = old_count + add_count

            old_name = existing[emp_id]["name"]
            if (not old_name) or (len(new_name) > len(old_name)):
                ws.cell(row=row_idx, column=2).value = new_name

            docs_set = existing[emp_id]["docs"]
            docs_set.add(doc_name)
            ws.cell(row=row_idx, column=4).value = ", ".join(sorted(docs_set))
        else:
            ws.append([emp_id, new_name, add_count, doc_name])

    wb.save(out_path)
    _mark_doc_submitted(export_name, doc_name)

    return {
        "status": "ok",
        "already_submitted": False,
        "file": str(out_path),
        "updated_employees": len(doc_counts),
        "document": doc_name,
        "export": export_name,
    }


# ==================================================
# OLD: Upload & process ONE PDF (optional)
# ==================================================
@app.post("/upload-pdf")
async def upload_pdf(file: UploadFile = File(...)):
    if file.content_type != "application/pdf":
        raise HTTPException(status_code=400, detail="Only PDF files are allowed")

    file_path = UPLOAD_DIR / file.filename
    file_path.parent.mkdir(exist_ok=True)

    import shutil
    with file_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    extracted_data = extract_fields_from_pdf(str(file_path))
    return {"filename": file.filename, "extracted_data": extracted_data}


# ==================================================
# OLD: Process ALL PDFs in local upload folder (optional)
# ==================================================
@app.post("/process-batch")
def process_batch():
    pdfs = list(UPLOAD_DIR.glob("*.pdf")) + list(UPLOAD_DIR.glob("*.PDF"))
    if not pdfs:
        raise HTTPException(status_code=400, detail="No PDFs found in upload folder")
    return process_pdf_batch(UPLOAD_DIR)