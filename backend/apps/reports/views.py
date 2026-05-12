from __future__ import annotations

from datetime import date
from io import BytesIO

from django.db.models import Count
from django.http import HttpResponse
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.documents.models import HPRecord, MonthBatch
from apps.documents.permissions import is_admin_user
from apps.documents.serializers import parse_month_label
from apps.employees.models import Department


def _assert_department_access(user, department_id: int) -> Department:
    dept = Department.objects.filter(pk=int(department_id)).first()
    if dept is None:
        raise ValidationError({"department": "Invalid department id."})
    if is_admin_user(user):
        return dept
    profile = getattr(user, "userprofile", None)
    user_dept = getattr(profile, "department", None)
    if user_dept is None or user_dept.id != dept.id:
        raise PermissionDenied("You do not have access to this department.")
    return dept


class MonthlyReportView(APIView):
    """GET /api/reports/monthly/?department=<id>&month=<MM-YYYY>"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        dept_id = request.query_params.get("department")
        month_label = request.query_params.get("month")
        if not dept_id:
            raise ValidationError({"department": "department is required."})
        if not month_label:
            raise ValidationError({"month": "month is required (MM-YYYY)."})

        dept = _assert_department_access(request.user, int(dept_id))
        month_date: date = parse_month_label(str(month_label))

        qs = HPRecord.objects.filter(department=dept, month_date=month_date).select_related(
            "employee", "department"
        )

        by_person = (
            qs.values(
                "employee__employee_id",
                "employee__full_name",
                "department__name",
                "employee__type",
            )
            .annotate(hp_count=Count("id"))
            .order_by("-hp_count", "employee__full_name")
        )

        items = [
            {
                "employee_id": r["employee__employee_id"],
                "full_name": r["employee__full_name"],
                "department_name": r["department__name"],
                "type": r["employee__type"],
                "hp_count": r["hp_count"],
            }
            for r in by_person
        ]

        summary = {
            "total_documents": qs.values("job_id").distinct().count(),
            "total_participations": qs.count(),
            "unique_persons": qs.values("employee_id").distinct().count(),
        }

        return Response({"summary": summary, "items": items})


class MonthlyReportTrendView(APIView):
    """GET /api/reports/monthly/trend/?department=<id>"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        dept_id = request.query_params.get("department")
        if not dept_id:
            raise ValidationError({"department": "department is required."})
        dept = _assert_department_access(request.user, int(dept_id))

        qs = HPRecord.objects.filter(department=dept).select_related("month_batch")
        rows = (
            qs.values("month_batch__month_label", "month_batch__month_date")
            .annotate(
                total_documents=Count("job_id", distinct=True),
                unique_persons=Count("employee_id", distinct=True),
            )
            .order_by("month_batch__month_date")
        )

        data = [
            {
                "month_label": r["month_batch__month_label"],
                "month_date": r["month_batch__month_date"],
                "total_documents": r["total_documents"],
                "unique_persons": r["unique_persons"],
            }
            for r in rows
        ]
        return Response(data)


class MonthlyReportExportView(APIView):
    """GET /api/reports/monthly/export/?department=<id>&month=<MM-YYYY>"""

    permission_classes = [IsAuthenticated]

    def get(self, request):
        dept_id = request.query_params.get("department")
        month_label = request.query_params.get("month")
        if not dept_id:
            raise ValidationError({"department": "department is required."})
        if not month_label:
            raise ValidationError({"month": "month is required (MM-YYYY)."})

        dept = _assert_department_access(request.user, int(dept_id))
        month_date: date = parse_month_label(str(month_label))

        qs = HPRecord.objects.filter(department=dept, month_date=month_date).select_related(
            "employee", "department"
        )
        by_person = (
            qs.values(
                "employee__employee_id",
                "employee__full_name",
                "department__name",
                "employee__type",
            )
            .annotate(hp_count=Count("id"))
            .order_by("-hp_count", "employee__full_name")
        )

        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        wb = Workbook()
        ws = wb.active
        ws.title = "Monthly Report"

        title = f"{dept.name} — {month_label}"
        ws["A1"] = title
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:E1")

        ws.append(["Employee ID", "Full Name", "Department", "Type", "HP Count"])
        for cell in ws[2]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for r in by_person:
            ws.append(
                [
                    r["employee__employee_id"],
                    r["employee__full_name"],
                    r["department__name"],
                    r["employee__type"],
                    r["hp_count"],
                ]
            )

        ws.append([])
        ws.append(
            [
                "Summary",
                "",
                "",
                "",
                "",
            ]
        )
        ws.append(["Total documents", qs.values("job_id").distinct().count(), "", "", ""])
        ws.append(["Total participations", qs.count(), "", "", ""])
        ws.append(["Unique persons", qs.values("employee_id").distinct().count(), "", "", ""])

        for col in ["A", "B", "C", "D", "E"]:
            ws.column_dimensions[col].width = 22

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        filename = f"monthly_report_{dept.code}_{month_label}.xlsx"
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename}"'
        return resp
