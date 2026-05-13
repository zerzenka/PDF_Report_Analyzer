from __future__ import annotations

import csv
import re
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand, CommandError
from openpyxl import load_workbook

from apps.employees.models import Department, Employee


def _department_code(name: str) -> str:
    base = re.sub(r"[^A-Za-z0-9]", "", (name or "").strip())[:20].upper() or "DEPT"
    code = base
    i = 1
    while Department.objects.filter(code=code).exists():
        suffix = str(i)
        code = (base[: max(1, 20 - len(suffix))] + suffix)[:20]
        i += 1
    return code


def _cell_int(value) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    return int(float(s))


def _normalize_csv_row(row: dict) -> dict[str, str]:
    """Strip header keys and values so 'Name ' / 'ID' still match."""
    out: dict[str, str] = {}
    for k, v in row.items():
        if k is None:
            continue
        key = str(k).strip()
        if not key:
            continue
        out[key] = "" if v is None else str(v).strip()
    return out


class Command(BaseCommand):
    help = (
        "Import employees from Excel (EmployeeID, EmployeeName, Department), "
        "or contractors from CSV (Name, ID) via --contractors."
    )

    def add_arguments(self, parser):
        default_path = Path(settings.BASE_DIR).parent / "employees.xlsx"
        parser.add_argument(
            "--file",
            type=str,
            default=str(default_path),
            help=f"Path to Excel file (default: {default_path})",
        )
        parser.add_argument(
            "--contractors",
            type=str,
            default=None,
            metavar="PATH",
            help="Path to contractors CSV with columns Name and ID.",
        )

    def handle(self, *args, **options):
        contractors_path = options.get("contractors")
        if contractors_path:
            self._import_contractors(Path(contractors_path).resolve())
            return

        path = Path(options["file"]).resolve()
        if not path.is_file():
            raise CommandError(f"File not found: {path}")

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            ws = wb.active
            rows = ws.iter_rows(min_row=1, values_only=True)
            header = next(rows, None)
            if not header or len(header) < 3:
                raise CommandError("Expected header row with at least 3 columns.")

            h0, h1, h2 = (str(c or "").strip() for c in header[:3])
            if h0.lower() != "employeeid" or h1.lower() != "employeename" or h2.lower() != "department":
                raise CommandError(
                    f"Unexpected headers: {header[:3]!r}. "
                    "Expected: EmployeeID, EmployeeName, Department"
                )

            created = 0
            updated = 0
            depts_created = 0

            for row in rows:
                if not row or all(c is None or str(c).strip() == "" for c in row[:3]):
                    continue
                emp_id_raw = row[0]
                name = str(row[1] or "").strip()
                dept_name = str(row[2] or "").strip()
                if not name or not dept_name:
                    continue

                eid = _cell_int(emp_id_raw)
                if eid is None:
                    continue
                emp_id_str = str(eid).zfill(6)

                dept, dept_was_created = Department.objects.get_or_create(
                    name=dept_name,
                    defaults={"code": _department_code(dept_name)},
                )
                if dept_was_created:
                    depts_created += 1

                _, was_created = Employee.objects.update_or_create(
                    employee_id=emp_id_str,
                    defaults={
                        "full_name": name,
                        "department": dept,
                        "type": "employee",
                        "is_active": True,
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1
        finally:
            wb.close()

        self.stdout.write(
            self.style.SUCCESS(
                f"Import complete: {created} created, {updated} updated, "
                f"{depts_created} departments created."
            )
        )

    def _import_contractors(self, path: Path) -> None:
        if not path.is_file():
            raise CommandError(f"File not found: {path}")

        created = 0
        updated = 0

        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames is None:
                raise CommandError("CSV has no header row.")
            headers = [(h or "").strip() for h in reader.fieldnames]
            if "Name" not in headers or "ID" not in headers:
                raise CommandError(
                    "CSV must include columns Name and ID. "
                    f"Found: {reader.fieldnames!r}"
                )

            for raw in reader:
                row = _normalize_csv_row(raw)
                name = row.get("Name", "").strip()
                id_raw = row.get("ID", "").strip()
                if not name and not id_raw:
                    continue
                if not id_raw:
                    self.stdout.write(
                        self.style.WARNING(f"Skipping row with empty ID: {raw!r}")
                    )
                    continue

                employee_id = str(id_raw).zfill(6)
                if not name:
                    self.stdout.write(
                        self.style.WARNING(
                            f"Skipping row with empty Name for ID {employee_id!r}"
                        )
                    )
                    continue

                _, was_created = Employee.objects.update_or_create(
                    employee_id=employee_id,
                    defaults={
                        "full_name": name,
                        "type": "contractor",
                        "department": None,
                        "is_active": True,
                    },
                )
                if was_created:
                    created += 1
                else:
                    updated += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"{created} contractors created, {updated} updated"
            )
        )
